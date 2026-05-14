#!/usr/bin/env python3
"""
Generate the SKU master-list spreadsheet for Jordan + Marvin to fill out.

Pre-populates with everything currently in lib/products.ts so they edit
existing rows instead of starting blank. Adds new rows for SKUs decided
on the call (Tirzepatide, AOD-9604, Epitalon, Glutathione, KPV, SS-31,
Wolverine blend, Glow blend). Yellow-highlights cells that need their
explicit input (price, exact dosage, tag membership, inventory grams).

Output: docs/viora-sku-template.xlsx
        ~/Downloads/viora-sku-template.xlsx (for easy email attach)

Run:    python3 scripts/generate-sku-template.py
"""
from __future__ import annotations
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "docs")
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "viora-sku-template.xlsx")
DOWNLOADS = os.path.expanduser("~/Downloads/viora-sku-template.xlsx")

# ─── Style tokens ───────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="284C3E")     # Viora green
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="ECF3EE")
NEEDS_INPUT_FILL = PatternFill("solid", fgColor="FFF4CC")  # soft yellow
WARN_FILL = PatternFill("solid", fgColor="FCE4E4")          # soft red
THIN = Side(style="thin", color="D0D7D2")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = CELL_BORDER


def style_data(cell, fill=None):
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = CELL_BORDER
    cell.font = Font(name="Calibri", size=10)
    if fill is not None:
        cell.fill = fill


# ─── SKU pre-population ─────────────────────────────────────────────
# Each row = one purchasable SKU (a specific compound + dosage combo).
# Pre-filled from lib/products.ts + lib/coas.ts + call decisions.
# "?" or empty = needs input from Jordan/Marv.

SKU_ROWS = [
    # SKU code      Display name              Compound              Dose    Price   Tags                                           Type    Grams   CAS                  Notes
    # ── Already in inventory per Jordan's call list ─────────────────
    ("VH3-R-10",   "VH3-R (Retatrutide)",    "Retatrutide",        "10mg", "",    "Best Seller, Weight Loss",                    "single", 27,   "2381089-83-2",      "Confirm price (Direct Peptides: $139 — go $5 below)"),
    ("VH3-R-30",   "VH3-R (Retatrutide)",    "Retatrutide",        "30mg", "",    "Best Seller, Weight Loss",                    "single", 27,   "2381089-83-2",      "Confirm price (Direct Peptides: $230 — go $5 below)"),
    ("VH2-T-10",   "VH2-T (Tirzepatide)",    "Tirzepatide",        "10mg", "",    "Weight Loss",                                 "single", 4,    "2023788-19-2",      "NEW — verify CAS. Currently labeled as Teduglutide in code, needs swap"),
    ("VH2-T-30",   "VH2-T (Tirzepatide)",    "Tirzepatide",        "30mg", "",    "Weight Loss",                                 "single", 4,    "2023788-19-2",      "NEW"),
    ("TES-10",     "Tesamorelin",            "Tesamorelin",        "10mg", "",    "Best Seller, Weight Loss",                    "single", 3,    "218949-48-5",       "Single dose per call"),
    ("BPC-157-10", "BPC-157",                "BPC-157",            "10mg", "",    "Longevity",                                   "single", 5,    "137525-51-0",       "Single dose per call"),
    ("TB500-10",   "TB-500",                 "TB-500",             "10mg", "",    "Longevity",                                   "single", 8,    "77591-33-4",        "Single dose per call"),
    ("GHK-100",    "GHK-Cu",                 "GHK-Cu",             "100mg","",    "Best Seller, Anti-Aging",                     "single", 150,  "89030-95-5",        "Single dose per call"),
    ("MOTS-10",    "MOTS-c",                 "MOTS-c",             "10mg", "",    "Best Seller, Energy",                         "single", 4,    "1627580-64-6",      "Single dose per call"),
    ("KPV-10",     "KPV",                    "KPV",                "10mg", "",    "Anti-Aging",                                  "single", 3,    "67727-97-3",        "Single dose per call. NEW — verify CAS"),
    ("SS31-",      "SS-31",                  "SS-31 (Elamipretide)","?",   "",    "Energy",                                      "single", 2,    "736992-21-5",       "DOSAGE?: Kyle will confirm. Direct Peptides has 50mg"),
    ("LR3-1",      "IGF-1 LR3",              "IGF-1 LR3",          "1mg",  "",    "Longevity",                                   "single", 3,    "946870-92-4",       "Direct Peptides: 1mg. Confirm we're carrying"),
    ("CJC-NoDAC-10","CJC-1295 No DAC",       "CJC-1295 (no DAC)",  "10mg", "",    "Longevity",                                   "single", 2,    "863288-34-0",       "No DAC per Marvin"),
    ("IPA-10",     "Ipamorelin",             "Ipamorelin",         "10mg", "",    "Longevity",                                   "single", 2,    "170851-70-4",       ""),
    ("NAD-1000",   "NAD+",                   "NAD+",               "1000mg","",   "Energy",                                      "single", 100,  "53-84-9",           "~80% of stock at 1000mg per Marvin"),
    ("NAD-500",    "NAD+",                   "NAD+",               "500mg","",    "Energy",                                      "single", 100,  "53-84-9",           "~20% of stock at 500mg per Marvin"),
    ("SEM-10",     "Semax",                  "Semax",              "10mg", "",    "Cognitive",                                   "single", 2,    "80714-61-0",        ""),
    ("SEL-10",     "Selank",                 "Selank",             "10mg", "",    "Cognitive",                                   "single", 2,    "129954-34-3",       ""),
    ("PT141-10",   "PT-141",                 "PT-141 (Bremelanotide)","10mg","",  "",                                            "single", 2,    "189691-06-3",       "TAG?: no obvious category — pick one"),
    ("AOD-5",      "AOD-9604",               "AOD-9604",           "5mg",  "",    "Weight Loss",                                 "single", 2,    "221231-10-3",       "NEW. Verify CAS"),
    ("EPI-10",     "Epitalon",               "Epitalon",           "10mg", "",    "Anti-Aging",                                  "single", 2,    "307297-39-8",       "Kyle to confirm dose. Direct Peptides has 10mg"),
    ("GLUT-500",   "Glutathione",            "Glutathione",        "500mg","",    "Anti-Aging",                                  "single", 50,   "70-18-8",           "500mg most common per Kyle. 1500mg variant possible"),

    # ── Blends (multi-peptide bundles in one vial) ──────────────────
    ("WOLV-10-10", "Wolverine Blend",        "BPC-157 + TB-500",   "10mg + 10mg","75",  "Longevity",                                  "blend",  "",   "",                  "Current price $75. Per call: both at 10mg, not 5+5"),
    ("CJCIPA-10-10","CJC-1295 + Ipamorelin", "CJC-1295 + Ipamorelin","10mg + 10mg","75","Longevity",                                "blend",  "",   "",                  "Current price $75. Both at 10mg per call"),
    ("GLOW-",      "Glow Blend",             "GHK-Cu + ?",         "?",    "85",   "Anti-Aging",                                  "blend",  "",   "",                  "Current price $85. DOSAGE? Need component breakdown from Marvin"),

    # ── Stacks (multiple separate vials sold as a bundle) ──────────
    # Each stack = customer gets N separate vials with their own batches.
    # Price column reflects current site price; needs your $5-below-DP cross-check.
    ("STK-PREMWL", "Viora Premium Weight Loss Stack", "VH3-R + Tesamorelin + MOTS-c", "3 vials", "388", "Best Seller, Weight Loss", "stack", "", "", "Current price $388. Confirm components + dose mix"),
    ("STK-WL",     "Viora Weight Loss Stack",         "Tesamorelin + VH3-R",         "2 vials", "303", "Weight Loss",              "stack", "", "", "Current price $303"),
    ("STK-MET",    "Viora Metabolic Stack",           "VH3-R + MOTS-c + NAD+",       "3 vials", "430", "Energy",                   "stack", "", "", "Current price $430. NOTE: 'Metabolic Health' tag removed — using 'Energy' instead. Confirm or kill?"),
    ("STK-GUT",    "Viora Gut Health Stack",          "BPC-157 + GLP-2 (Teduglutide)","2 vials","265", "Longevity",                "stack", "", "", "Current price $265. ⚠️ Includes Teduglutide which we're removing — replace with what?"),
    ("STK-LNG",    "Viora Longevity Stack",           "NAD+ + MOTS-c + GHK-Cu",      "3 vials", "189", "Longevity, Anti-Aging",    "stack", "", "", "Current price $189"),
    ("STK-HORM",   "Viora Hormone Signaling Stack",   "Tesamorelin + Ipamorelin + IGF-1 LR3","3 vials","102","REMOVE?",            "stack", "", "", "Current price $102. ⚠️ 'Hormone' tag removed per call. Kill this stack or rename?"),
    ("STK-MOOD",   "Viora Mood & Balance Stack",      "Selank + Semax + Oxytocin",   "3 vials", "70",  "Cognitive",                "stack", "", "", "Current price $70. ⚠️ Contains Oxytocin (REMOVE candidate). Keep without Oxytocin? Or drop entire stack?"),
    ("STK-INT",    "Viora Intimacy Research Stack",   "PT-141 + Oxytocin",           "2 vials", "50",  "REMOVE?",                  "stack", "", "", "Current price $50. ⚠️ Contains Oxytocin (REMOVE candidate). Kill?"),
    ("STK-CEO",    "Viora CEO Stack",                 "NAD+ + Semax + Selank",       "3 vials", "100", "Cognitive, Energy",        "stack", "", "", "Current price $100. Confirm — performance positioning"),

    # ── SKUs to consider — were in old code but NOT confirmed for launch ─
    ("REMOVE-1",   "GLP-2 (Teduglutide)",    "Teduglutide",        "5mg",  "",    "REMOVE",                                      "single", "",   "223132-37-4",       "Decision: replace with Tirzepatide (VH2-T)? Or keep both?"),
    ("REMOVE-2",   "Oxytocin",               "Oxytocin",           "2mg",  "",    "REMOVE?",                                     "single", "",   "50-56-6",           "Was in catalog — keep or drop? Affects 3 stacks (Mood, Intimacy)"),
    ("REMOVE-3",   "DSIP",                   "DSIP",               "",     "",    "REMOVE",                                      "single", "",   "62568-57-4",        "Confirmed NOT in inventory per Jordan — remove"),
    ("REMOVE-4",   "5-Amino 1MQ",            "5-Amino 1MQ",        "",     "",    "REMOVE",                                      "single", "",   "",                  "Confirmed NOT in inventory per Jordan — remove"),
]

HEADERS = [
    "SKU code",
    "Display name (shown on site/label)",
    "Compound (chemistry)",
    "Dosage per vial",
    "Retail price (USD, no $)",
    "Tags (comma-separated)",
    "Type",
    "Grams in inventory",
    "CAS #",
    "Notes / questions",
]

# Cells that need input get yellow highlight
NEEDS_INPUT_COLS = {"Retail price (USD, no $)"}


# ─── Write the workbook ─────────────────────────────────────────────
wb = Workbook()

# Sheet 1: SKUs (the main editable list)
ws = wb.active
ws.title = "SKUs"
ws.sheet_view.zoomScale = 110

# Title row
ws.merge_cells("A1:J1")
title = ws["A1"]
title.value = "Viora Healthcare — SKU Master List"
title.font = Font(name="Calibri", bold=True, size=16, color="284C3E")
title.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("A2:J2")
sub = ws["A2"]
sub.value = "Fill in price, confirm dosages, fix any cells marked NEW or with question marks. Yellow cells need input. Red cells are REMOVE candidates — delete the row if confirmed gone."
sub.font = Font(name="Calibri", italic=True, size=10, color="5A6A5E")
sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 36

# Header row (row 4)
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    style_header(cell)
ws.row_dimensions[4].height = 32

# Data rows (start row 5)
for row_idx, row_data in enumerate(SKU_ROWS, start=5):
    sku_code = row_data[0]
    is_remove = sku_code.startswith("REMOVE-")

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

        if is_remove:
            style_data(cell, fill=WARN_FILL)
        elif HEADERS[col_idx - 1] in NEEDS_INPUT_COLS:
            style_data(cell, fill=NEEDS_INPUT_FILL)
        else:
            style_data(cell)

# Column widths tuned to content
COL_WIDTHS = [14, 28, 22, 14, 11, 32, 9, 9, 17, 50]
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top
ws.freeze_panes = "A5"


# Sheet 2: Allowed Tags reference
ref = wb.create_sheet("Allowed Tags")
ref.column_dimensions["A"].width = 24
ref.column_dimensions["B"].width = 50

ref["A1"] = "Allowed Tag"
ref["B1"] = "Description / What goes here"
for cell in [ref["A1"], ref["B1"]]:
    style_header(cell)

TAGS = [
    ("Best Seller", "Auto-rotates with top-selling SKUs once we have sales data. Until then: Reta, MOTS-c, Tessa, GHK-Cu, Wolverine"),
    ("Weight Loss", "VH2-T, VH3-R, Tesamorelin, AOD-9604"),
    ("Anti-Aging", "GHK-Cu, KPV, Glutathione, Epitalon, Glow"),
    ("Longevity", "TB-500, BPC-157, CJC-1295, Ipamorelin, IGF-1 LR3"),
    ("Cognitive", "Semax, Selank"),
    ("Energy", "NAD+, MOTS-c, SS-31"),
    ("REMOVE", "Mark a row REMOVE if we are NOT selling this SKU at launch (we'll delete from the site)"),
]
for i, (tag, desc) in enumerate(TAGS, start=2):
    ref.cell(row=i, column=1, value=tag)
    ref.cell(row=i, column=2, value=desc)
    for col in (1, 2):
        style_data(ref.cell(row=i, column=col))

# Note: removed tags from the call
ref.merge_cells(f"A{len(TAGS) + 3}:B{len(TAGS) + 3}")
removed_note = ref.cell(row=len(TAGS) + 3, column=1)
removed_note.value = "NOTE: We removed 'Metabolic Health' and 'Hormone' categories per the call — no hormone products being sold."
removed_note.font = Font(name="Calibri", italic=True, size=10, color="5A6A5E")
removed_note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# Sheet 3: Inventory reference (from Jordan's call readout)
inv = wb.create_sheet("Inventory (Jordan's list)")
inv.column_dimensions["A"].width = 26
inv.column_dimensions["B"].width = 12
inv.column_dimensions["C"].width = 40

inv["A1"] = "Compound"
inv["B1"] = "Grams ordered"
inv["C1"] = "Notes"
for col in "ABC":
    style_header(inv[f"{col}1"])

INVENTORY = [
    ("Retatrutide",       27,  "main weight-loss SKU"),
    ("Tesamorelin",       3,   ""),
    ("BPC-157",           5,   ""),
    ("TB-500",            8,   ""),
    ("GHK-Cu",            150, "biggest stock — anti-aging push"),
    ("MOTS-c",            4,   ""),
    ("KPV",               3,   ""),
    ("SS-31",             2,   ""),
    ("IGF-1 LR3",         3,   ""),
    ("CJC-1295 (no DAC)", 2,   ""),
    ("NAD+",              100, "split 80/20 between 1000mg and 500mg"),
    ("Semax",             2,   ""),
    ("Selank",            2,   ""),
    ("PT-141",            2,   ""),
    ("Tirzepatide",       4,   "NEW — replacing GLP-2-T (Teduglutide) in catalog"),
    ("AOD-9604",          2,   ""),
    ("Epitalon",          2,   ""),
    ("Glutathione",       50,  ""),
]
for i, (compound, grams, notes) in enumerate(INVENTORY, start=2):
    inv.cell(row=i, column=1, value=compound)
    inv.cell(row=i, column=2, value=grams)
    inv.cell(row=i, column=3, value=notes)
    for col in (1, 2, 3):
        style_data(inv.cell(row=i, column=col))

inv.merge_cells(f"A{len(INVENTORY) + 3}:C{len(INVENTORY) + 3}")
inv_note = inv.cell(row=len(INVENTORY) + 3, column=1)
inv_note.value = "Source: Jordan's readout on the team call. NOT carried this run: DSIP, IGF (alternate forms), 5-Amino 1MQ."
inv_note.font = Font(name="Calibri", italic=True, size=10, color="5A6A5E")
inv_note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# Sheet 4: How to fill out
instr = wb.create_sheet("How to fill out")
instr.column_dimensions["A"].width = 90
instr["A1"] = "How to fill this out"
style_header(instr["A1"])
instr.row_dimensions[1].height = 28

INSTRUCTIONS = [
    "",
    "1. SKU code — keep the VH-style codes we agreed on (VH3-R-10, VH2-T-30, etc). Each dosage is its own SKU.",
    "",
    "2. Display name — what shows up on the storefront, vial label, COA. Format: 'VH3-R (Retatrutide)' or just 'Tesamorelin' for non-VH SKUs.",
    "",
    "3. Compound — the actual chemistry name (Retatrutide, Tirzepatide, BPC-157). Single source of truth for the COA generator.",
    "",
    "4. Dosage — milligrams per vial. Always include 'mg' suffix. Examples: 10mg, 30mg, 100mg, 500mg, 10mg + 10mg (for blends).",
    "",
    "5. Retail price (highlighted yellow) — what the customer pays. Per Marvin's call: $5 below Direct Peptides for every SKU. If Direct sells Retatrutide 10mg at $139, we sell at $134.",
    "",
    "6. Tags — pick from the Allowed Tags sheet. Multiple OK, comma-separated. Examples: 'Best Seller, Weight Loss' or 'Anti-Aging'.",
    "",
    "7. Type — single | blend | stack",
    "",
    "8. Grams in inventory — from Jordan's readout on the call. Helps sanity-check the dosage math (10mg vials from 27g = 2700 vials possible).",
    "",
    "9. CAS # — chemistry ID, goes on the COA. We already have most of these. Verify NEW ones (Tirzepatide, KPV, AOD, etc).",
    "",
    "10. Notes — anything else: special handling, restrictions, marketing positioning, etc.",
    "",
    "RED rows are REMOVE candidates — confirm by leaving them red, OR delete the row entirely if we're definitely dropping that SKU.",
    "",
    "BLENDS vs STACKS:",
    "   • Blend = one vial containing multiple peptides (Wolverine = BPC + TB-500 in one vial)",
    "   • Stack = multiple separate vials sold as a bundle (Premium Weight Loss = 3 separate vials)",
    "",
    "STACK COMPOSITION QUESTIONS to resolve:",
    "   • Gut Health Stack currently includes Teduglutide which we're removing — pick a replacement or kill the stack",
    "   • Hormone Signaling Stack — 'Hormone' tag was removed on the call. Kill the stack or just rename it?",
    "   • Mood & Balance + Intimacy Research stacks both contain Oxytocin (REMOVE candidate). Kill the stacks or restructure without Oxytocin?",
    "",
    "Send the completed sheet back to Farris. He'll bulk-update the storefront, COA registry, and label generation from it in one batch.",
]

for i, line in enumerate(INSTRUCTIONS, start=2):
    cell = instr.cell(row=i, column=1, value=line)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.font = Font(name="Calibri", size=11)
    if line and not line.startswith(" ") and (line[0].isdigit() or line.startswith("RED") or line.startswith("Send")):
        cell.font = Font(name="Calibri", size=11, bold=True, color="284C3E")
    instr.row_dimensions[i].height = 18

# Reorder sheets so SKUs is first, instructions second
wb.move_sheet("How to fill out", offset=-3)
wb.move_sheet("Allowed Tags", offset=-2)
wb.move_sheet("Inventory (Jordan's list)", offset=-1)

# Save
wb.save(OUT)
shutil.copyfile(OUT, DOWNLOADS)

print(f"✓ Wrote {OUT}")
print(f"✓ Wrote {DOWNLOADS} (for email attach)")
print()
print(f"  Sheets: SKUs ({len(SKU_ROWS)} rows pre-filled) · How to fill out · Allowed Tags · Inventory")
